from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import *
from django.views.decorators.csrf import csrf_protect
from datetime import date, datetime, timedelta
from django.contrib import messages
from django.utils.timezone import make_aware
from django.core.mail import send_mail, EmailMessage
from django.conf import settings
import xlrd
import xlwt
from braces import views
from django.contrib.auth.models import User
from django.db.models import Q
from .forms import ProjectForm
from .forms import TranscriptForm, TranscriptForm2, AddressForm, GradesheetForm, GradesheetAddressForm, GradesheetForm2, GradesheetForm3, GradesheetCopiesForm
import PyPDF2
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import openpyxl
import pandas as pd
from openpyxl import load_workbook


module_dir = os.path.dirname(__file__)  # get current directory

file_path = os.path.join(module_dir, 'tr.pdf')
pdfFileObj = open(file_path, 'rb')
pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
global n
n = pdfReader.numPages

fromaddr = "ish2siddh@gmail.com"


def rank(a, b, c, d, e):
    if(a):
        return 1
    if(b):
        return 4
    if(c):
        return 0
    if(d):
        return 2
    if(e):
        return 3


def funct_to_sort():

    all_objects = Transcript.objects.all()

    projects = sorted(all_objects, key=lambda t: rank(
        t.approved, t.disapproved, t.inprocess, t.posted, t.printed))

    # context= {'all_objects': all_objects,
    #           'option':1,
    #           'hfuser': Arc.objects.get(user=request.user),
    #           'project': projects,
    #           }

    return projects


def funct_to_sort_2():

    all_objects = Gradesheet.objects.all()

    projects = sorted(all_objects, key=lambda t: rank(
        t.approved, t.disapproved, t.inprocess, t.posted, t.printed))

    # context= {'all_objects': all_objects,
    #           'option':1,
    #           'hfuser': Arc.objects.get(user=request.user),
    #           'project': projects,
    #           }

    return projects


def is_hod(user):
    return False if not Arc.objects.filter(user=user) else True


@login_required
@user_passes_test(is_hod)
def hod(request):
    hod = Arc.objects.get(user=request.user)
    projects = Transcript.objects.all()
    context = {
        'option': 1,
        'hfuser': hod,
        'project': funct_to_sort(),
    }
    postContext = {
        'projects': funct_to_sort
    }
    return render(request, "hod.html", dict(context, **postContext))


# def is_faculty(user):
#      return False if not Faculty.objects.filter(user=user) else True
def index(request):
    # send_mail('Hello from AUGSD', 'Hello there, this is a message!',
    #           settings.EMAIL_HOST_USER, [toaddr], fail_silently=False)
    return render(request, 'home1.html', {})


def login_success(request):
    return HttpResponse("Success!")

# @login_required
# @user_passes_test(is_faculty)
# def faculty(request):
#     faculty = Faculty.objects.filter(user=request.user)
#     context ={
#         'faculty' : faculty,
#     }
#     return render(request, "faculty.html",)


@csrf_protect
def loginform(request):

    if request.user.is_authenticated:
        if request.user.is_staff:
            return redirect('/admin')
        if Arc.objects.filter(user=request.user):
            return redirect('/arc')
        # if Faculty.objects.filter(user=request.user):
        #     return redirect('/project')
        return redirect('/')

    if request.POST:
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            if request.user.is_staff:
                return redirect('/admin')
            if Arc.objects.filter(user=request.user):
                return redirect('/arc')
            # if Faculty.objects.filter(user=request.user):
            #     return redirect('/project')
        else:
            messages.add_message(
                request, messages.INFO,  "Incorrect username or password", extra_tags='red')
            print('Not able to authenticate')

    return render(request, "sign-in.html", {})


@login_required
def logoutform(request):
    logout(request)
    return render(request, "logout.html", {})


@login_required
@user_passes_test(is_hod)
def arcGradesheet(request):
    hod = Arc.objects.get(user=request.user)
    projects = Gradesheet.objects.all()
    context = {
        'option': 1,
        'hfuser': hod,
        'project': funct_to_sort_2,
    }
    postContext = {
        'projects': funct_to_sort_2
    }
    return render(request, "arc-gradesheet.html", dict(context, **postContext))


@login_required
@user_passes_test(is_hod)
def upload(request):
    if "GET" == request.method:
        return render(request, 'upload.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size
        print('updating excel')
        dfs = pd.read_excel(excel_file, sheet_name='Sheet1')

        dfo = pd.read_excel('first_excel.xlsx', sheet_name='Sheet1')

        for index, row in dfo.iterrows():
            ref = row['Bank Reference No']
            paidAmt = row['Amount']
            # print(ref)
            # `print(ref,paidAmt)
            t = Transcript.objects.filter(refNo=ref)
            g = Gradesheet.objects.filter(refNo=ref)
            if (len(t) == 0 and len(g) == 0):
                print("lol1")
                # dfo = dfo.append(row)
            elif (len(t) != 0):
                if (paidAmt == t[0].amtcal):
                    # print("sahi")
                    dfo.drop(index, inplace=True)
                    t[0].paymentDisapproved = False
                    t[0].paymentApproved = True
                    t[0].paidAmount = paidAmt
                    t[0].save()
                else:
                    # print("galat")
                    dfo.drop(index, inplace=True)
                    t[0].paymentDisapproved = True
                    t[0].paidAmount = paidAmt
                    t[0].save()
            else:
                if (paidAmt == g[0].amtcal):
                    # print("sahi")
                    dfo.drop(index, inplace=True)
                    g[0].paymentDisapproved = False
                    g[0].paymentApproved = True
                    g[0].paidAmount = paidAmt
                    g[0].save()
                else:
                    # print("galat")
                    dfo.drop(index, inplace=True)
                    g[0].paymentDisapproved = True
                    g[0].paidAmount = paidAmt
                    g[0].save()

        # print(dfs)
        for index, row in dfs.iterrows():
            ref = row['Bank Reference No']
            paidAmt = row['Amount']
            t = Transcript.objects.filter(refNo=ref)
            g = Gradesheet.objects.filter(refNo=ref)
            # print(ref)
            if (len(t) == 0 and len(g) == 0):
                print("lol")
                ####
                # check for existing ref no in dfo before appending for multiple excel
                # uploads of same payment
                ####

                # dfs[index].to_excel(writer,sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, index = False,header= False)
            elif (len(t) != 0):
                if (paidAmt == t[0].amtcal):
                    print("sahi")
                    t[0].paymentDisapproved = False
                    t[0].paymentApproved = True
                    t[0].paidAmount = paidAmt
                    t[0].save()
                    dfo = dfo.append(row)
                else:
                    print("galat")
                    t[0].paymentDisapproved = True
                    t[0].paidAmount = paidAmt
                    t[0].save()
            else:
                if (paidAmt == g[0].amtcal):
                    print("sahi")
                    g[0].paymentDisapproved = False
                    g[0].paymentApproved = True
                    g[0].paidAmount = paidAmt
                    g[0].save()
                    dfo = dfo.append(row)
                else:
                    print("galat")
                    g[0].paymentDisapproved = True
                    g[0].paidAmount = paidAmt
                    g[0].save()

        # for row in worksheet.iter_rows():
        #     row_data = list()
        #     for cell in row:
        #         row_data.append(str(cell.value))
        #     excel_data.append(row_data)

        # print(dfo.head())

        # refs = Transcript.objects.get(id=project)
        dfo.to_excel("first_excel.xlsx")
        # print("#############")
        # t = Transcript.objects.get(refNo="ed")
        # print(t.bitsID, t.amtcal)
        return render(request, 'upload.html', {})


@login_required
@user_passes_test(is_hod)
def transcript_upload(request):
    if "GET" == request.method:
        return render(request, 'transcript_upload.html', {})
    else:
        transcript_pdf_file = request.FILES["transcript_pdf_file"]

        # you may put validations here to check extension or file size
        print('updating transcript pdf')

        # module_dir = os.path.dirname(__file__)  # get current directory
        # print(module_dir)
        # file_path_2 = os.path.join(module_dir, 'trial.pdf')
        # pdfFileObj_2 = open(transcript_pdf_file, 'rb')
        # pdfReader_2 = PyPDF2.PdfFileReader(pdfFileObj_2)
        # n = pdfReader_2.numPages
        # dfs = pd.read_excel(transcript_pdf_file, sheet_name='Sheet1')
        global n
        print("num of pages before uplaoding :", n)
        pdfFileObj = transcript_pdf_file
        pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
        num = pdfReader.numPages

        # i = 0
        # content = []
        # while (i<NumPages):
        #     text = pdfReader_2.getPage(i)
        #     content.append(text.extractText())
        #     i +=1
        pages = []
        print("um of pages after uploading : ", num)
        for i in range(num):
            pageObj = pdfReader.getPage(i)
            txt = pageObj.extractText()
            pages.append(i)
        pdfWriter = PyPDF2.PdfFileWriter()

        for i in pages:
            pageObj = pdfReader.getPage(i)
            pdfWriter.addPage(pageObj)

        outputFile = open('main/tr' + ".pdf", "wb")
        pdfWriter.write(outputFile)
        pdfFileObj.close()
        outputFile.close()
        n = num
        return render(request, 'transcript_upload.html', {})


# @login_required
# @user_passes_test(is_faculty)
# def faculty(request):
#     faculty = Faculty.objects.filter(user=request.user)
#     return render(request, "hostelsuperintendent.html", context)

# #@login_required
# #@user_passes_test(is_faculty)
# def project(request):
#     faculty = 'abd'#Faculty.objects.get(user=request.user)
#     form = ProjectForm()
#     web = 'base.html'
#     if request.user is not None:
#             login(request, request.user)
#             if Arc.objects.filter(user=request.user):
#                 web ='hodbase.html'
#     context = {
#         'option' : 0,
#         'hfuser': faculty,
#         'form': form,
#         'web':web
#     }

#     projectContext = {
#         'projects': 'abcd'#Project.objects.filter(faculty=faculty),
#     }

#     if request.POST:
#         form = ProjectForm(request.POST)
#         if form.is_valid():
#             projectform = form.save(commit=False)
#             projectform.faculty = faculty
#             # print(request.POST.get('consent'))
#             projectform.save()

#             context = {
#                 'option': 1,
#             }
#         else:
#             context = {
#                 'option': 2,
#                 'form': form
#             }
#             print(form.errors)
#     return render(request, "project.html", dict(context, **projectContext))


@login_required
@user_passes_test(is_hod)
def hodprojectapprove(request, project):
    project = Transcript.objects.get(id=project)
    hod = Arc.objects.get(user=request.user)
    print(file_path)
    # leaves = Leave.objects.filter(student=leave.student)
    # project.amtcal=200*project.origTranscript+100*project.dupTranscripts
    context = {
        'option': 2,
        'hod': hod,
        'project': project,
        # 'faculty': project.faculty
    }

    if request.POST:
        approved = request.POST.getlist('group1')
        comment = request.POST.getlist('group2')[0]
        print(comment, approved)
        # str(request.POST.getlist('bitsID')[0])#'2015A3PS0248G'
        bitsid = project.bitsID
        # str(request.POST.getlist('corrEmail')[0])#"f20160052@goa.bits-pilani.ac.in"
        toaddr = project.corrEmail
        # print(bitsid,'addr')
        msg = MIMEMultipart()
        msg['From'] = fromaddr
        msg['To'] = toaddr
        msg['Subject'] = 'Official Transcript'
        # print(msg)

        if '1' in approved:
            project.approved = True
            project.disapproved = False
            project.inprocess = False
            project.posted = False
            project.printed = False

            body = "Hello " + project.name + \
                ", \nThis is an auto - generated mail from ARC after approval of your transcript application.\n"

            print(body)

            if (len(comment) > 0):
                body += ("Remarks: \n" + comment)

            pages = []
            print(n)
            for i in range(n):
                pageObj = pdfReader.getPage(i)
                txt = pageObj.extractText()
                if bitsid in txt:
                    pages.append(i)
            pdfWriter = PyPDF2.PdfFileWriter()

            for i in pages:
                pageObj = pdfReader.getPage(i)
                pdfWriter.addPage(pageObj)

            outputFile = open(bitsid + ".pdf", "wb")
            pdfWriter.write(outputFile)
            pdfFileObj.close()
            outputFile.close()
            filename = bitsid + ".pdf"
            attachment = open(filename, "rb")

            email = EmailMessage('Transcript soft copy ready',
                                 'Greetings from AUGSD, the soft copy of your transcript is ready! PFA', settings.EMAIL_HOST_USER, [toaddr])
            # email.content_subtype = 'html'

            # file = open("README.md", "r")
            # email.attach("README.md", file.read(), 'text/plain')
            email.attach_file(filename)

            email.send()

            # p = MIMEBase('application', 'octet-stream')
            # p.set_payload((attachment).read())
            # encoders.encode_base64(p)
            # p.add_header('Content-Disposition',
            #              "attachment; filename= %s" % filename)
            # msg.attach(p)
        elif '2' in approved:
            project.disapproved = True
            project.approved = False
            project.inprocess = False
            project.posted = False
            project.printed = False
            body = "Hello " + project.name + \
                ", \nThis is an auto - generated mail from ARC after disapproval of your transcript application.\n"
            print(body)
            if (len(comment) > 0):
                body += ("Remarks: " + comment)

        elif '3' in approved:
            project.inprocess = False
            project.approved = False
            project.disapproved = False
            project.posted = False
            project.printed = True

            body = "Hello " + project.name + \
                ", \nThis is an auto - generated mail from ARC after printing your transcripts.\n"

            if (len(comment) > 0):
                body += ("Remarks: " + comment)

        elif '4' in approved:
            project.inprocess = False
            project.approved = False
            project.disapproved = False
            project.posted = True
            project.printed = False

            body = "Hello " + project.name + \
                ", \nThis is an auto - generated mail from ARC after posting your transcripts.\n"

            if (len(comment) > 0):
                body += ("Remarks: " + comment)

        else:
            project.inprocess = True
            project.approved = False
            project.disapproved = False
            project.posted = False
            project.printed = False
            project.save()
            return redirect('arc')

        '''msg.attach(MIMEText(body, 'plain'))
        s = smtplib.SMTP('smtp.gmail.com', 587)
        s.starttls()
        s.login(fromaddr, "transcript1@2")
        text = msg.as_string()
        s.sendmail(fromaddr, toaddr, text)
        s.quit() '''
        if(project.approved == False):
            send_mail('Official Transcript', body,
                      settings.EMAIL_HOST_USER, [toaddr], fail_silently=False)

        project.save()
        return redirect('arc')

    return render(request, 'hod.html', context)


@login_required
@user_passes_test(is_hod)
def arcgradesheetapprove(request, project):
    project = Gradesheet.objects.get(id=project)
    hod = Arc.objects.get(user=request.user)

    # leaves = Leave.objects.filter(student=leave.student)
    # project.amtcal=200*project.origTranscript+100*project.dupTranscripts
    context = {
        'option': 2,
        'hod': hod,
        'project': project,
        # 'faculty': project.faculty
    }

    if request.POST:
        approved = request.POST.getlist('group1')
        comment = request.POST.getlist('group2')[0]
        print(comment)
        # str(request.POST.getlist('bitsID')[0])#'2015A3PS0248G'
        bitsid = project.bitsID
        # str(request.POST.getlist('corrEmail')[0])#"f20160052@goa.bits-pilani.ac.in"
        toaddr = project.corrEmail
        msg = MIMEMultipart()
        msg['From'] = fromaddr
        msg['To'] = toaddr
        msg['Subject'] = 'Gradesheet Application'
        print(toaddr)
        if '1' in approved:
            project.approved = True
            project.disapproved = False
            project.inprocess = False
            project.posted = False
            project.printed = False

            body = "Hello " + project.name + \
                ", \nThis is an auto - generated mail from ARC after approval of your gradesheet application.\n"

            if (len(comment) > 0):
                body += ("Remarks: \n" + comment)

            # pages = []

            # for i in range(n):
            #     pageObj = pdfReader.getPage(i)
            #     txt = pageObj.extractText()
            #     if bitsid in txt:
            #         pages.append(i)
            # pdfWriter = PyPDF2.PdfFileWriter()

            # for i in pages:
            #     pageObj = pdfReader.getPage(i)
            #     pdfWriter.addPage(pageObj)

            # outputFile = open(bitsid + ".pdf", "wb")
            # pdfWriter.write(outputFile)
            # pdfFileObj.close()
            # outputFile.close()
            # filename = bitsid + ".pdf"
            # attachment = open(filename, "rb")
            # p = MIMEBase('application', 'octet-stream')
            # p.set_payload((attachment).read())
            # encoders.encode_base64(p)
            # p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
            # msg.attach(p)

        elif '2' in approved:
            project.disapproved = True
            project.approved = False
            project.inprocess = False
            project.posted = False
            project.printed = False
            body = "Hello " + project.name + \
                ", \nThis is an auto - generated mail from ARC after disapproval of your gradesheet application.\n"

            if (len(comment) > 0):
                body += ("Remarks: " + comment)

        elif '3' in approved:
            project.inprocess = False
            project.approved = False
            project.disapproved = False
            project.posted = False
            project.printed = True

            body = "Hello " + project.name + \
                ", \nThis is an auto - generated mail from ARC after printing your gradesheets.\n"

            if (len(comment) > 0):
                body += ("Remarks: " + comment)

        elif '4' in approved:
            project.inprocess = False
            project.approved = False
            project.disapproved = False
            project.posted = True
            project.printed = False

            body = "Hello " + project.name + \
                ", \nThis is an auto - generated mail from ARC after posting your gradesheets.\n"

            if (len(comment) > 0):
                body += ("Remarks: " + comment)

        else:
            project.inprocess = True
            project.approved = False
            project.disapproved = False
            project.posted = False
            project.printed = False
            project.save()
            return redirect('hod')

        ''' msg.attach(MIMEText(body, 'plain'))
        s = smtplib.SMTP('smtp.gmail.com', 587)
        s.starttls()
        s.login(fromaddr, "transcript1@2")
        text = msg.as_string()
        s.sendmail(fromaddr, toaddr, text)
        s.quit() '''

        send_mail('Gradesheet Application', body,
                  settings.EMAIL_HOST_USER, [toaddr], fail_silently=False)

        project.save()
        return redirect('arc-gradesheet')

    return render(request, "arc-gradesheet.html", context)


transcriptsDict = {}


def transcript2(request, bits_id):

    if bits_id not in transcriptsDict:
        return redirect('/transcript')

    transcriptform = transcriptsDict.get(bits_id)
    if (transcriptform.gradStatus == 'continuing'):
        transcriptform.amtcal = int(
            transcriptform.origTranscript)*200 + int(transcriptform.dupTranscripts)*100
    else:
        transcriptform.amtcal = int(transcriptform.dupTranscripts)*500

    if (int(transcriptform.sealedCopies) > 0):
        transcriptform.amtcal += 40+20*int(transcriptform.sealedCopies)

    if (transcriptform.postalAddressLocation == 'Within India'):
        transcriptform.amtcal += 100
    elif (transcriptform.postalAddressLocation == 'Outside India'):
        transcriptform.amtcal += 3000

    amtcal = transcriptform.amtcal
    no_addr = transcriptform.sealedCopies
    form = TranscriptForm2(extra=str(no_addr))
    form.amtcal = transcriptform.amtcal
    context = {
        'option': 0,
        'amtcal': amtcal,
        'form': form
    }

    transcriptContext = {
        'transcripts': Transcript.objects.all(),
    }

    if request.POST:
        form = TranscriptForm2(request.POST, extra=str(no_addr))
        if (request.POST.get('approved')):
            print('***** APPROVED *****')
        if form.is_valid():
            transcriptform.refNo = request.POST.get('refNo')
            transcriptform.save()

            for i in range(int(no_addr)):
                addrForm = AddressForm()
                addrForm = addrForm.save(commit=False)
                addrForm.addr = request.POST.get('Sealed address %s' % i)
                addrForm.transcript = Transcript.objects.get(
                    form_no=transcriptform.form_no)
                addrForm.save()

            context = {
                'option': 1,
                'amtcal': amtcal,
                'refNo': request.POST.get('refNo'),
                'form': transcriptform,
                'soft_hard': transcriptform.soft_hard_copy,
            }
            del transcriptsDict[bits_id]
        else:
            context = {
                'option': 2,
                'amtcal': amtcal,
                'form': form
            }
            print(form.errors)

    return render(request, "transcript2.html", dict(context, **transcriptContext))


def transcript(request):
    form = TranscriptForm()
    context = {
        'option': 0,
        'form': form
    }

    transcriptContext = {
        'transcripts': Transcript.objects.all(),
    }

    if request.POST:
        form = TranscriptForm(request.POST)
        if (request.POST.get('approved')):
            print('***** APPROVED *****')
        if form.is_valid():

            transcriptform = form.save(commit=False)
            transcriptform.corrId = request.POST.get('bitsID')
            transcriptform.corrname = request.POST.get('name')
            transcriptform.corrhostel = request.POST.get('hostel')
            transcriptform.corrroomNo = request.POST.get('roomNo')
            transcriptform.ps2Station = request.POST.get('ps2Station')
            transcriptform.email = request.POST.get('corrEmail')
            transcriptform.phone_number = request.POST.get('phone_number')
            transcriptform.origTranscript = request.POST.get('origTranscript')
            transcriptform.dupTranscripts = request.POST.get('dupTranscripts')
            transcriptform.forwardingLetters = request.POST.get(
                'forwardingLetters')
            transcriptform.sealedCopies = request.POST.get('sealedCopies')
            transcriptform.postalAddress = request.POST.get('postalAddress')
            transcriptform.postalAddressLocation = request.POST.get(
                'postalAddressLocation')
            transcriptform.soft_hard_copy = request.POST.get('soft_hard_copy')

            transcriptsDict[transcriptform.corrId] = transcriptform
            # transcriptform.save()
            bitsid = transcriptform.corrId
            toaddr = transcriptform.email

            context = {
                'option': 1,
                'dupTranscripts': request.POST.get('dupTranscripts'),
                'form': form,
            }

            pages = []
            print(n)
            for i in range(n):
                pageObj = pdfReader.getPage(i)
                txt = pageObj.extractText()
                if bitsid in txt:
                    pages.append(i)
            pdfWriter = PyPDF2.PdfFileWriter()

            for i in pages:
                pageObj = pdfReader.getPage(i)
                pdfWriter.addPage(pageObj)

            outputFile = open(bitsid + ".pdf", "wb")
            pdfWriter.write(outputFile)
            pdfFileObj.close()
            outputFile.close()
            filename = bitsid + ".pdf"

            if(transcriptform.soft_hard_copy == 'Soft Copy'):
                transcriptform.approved = True
                transcriptform.inprocess = False
                transcriptform.paymentInprocess = False
                transcriptform.paymentApproved = True
                # transcriptform.save()
                context = {
                    'option': 1,
                    'amtcal': 0,
                    'refNo': 'NA',
                    'form': transcriptform,
                }
                del transcriptsDict[transcriptform.corrId]
                if request.POST:
                    transcriptform.refNo = 'NA'
                    transcriptform.amtcal = 0
                    transcriptform.save()

                    email = EmailMessage('Transcript soft copy ready',
                                         'Greetings from AUGSD, the soft copy of your transcript is ready! PFA', settings.EMAIL_HOST_USER, [toaddr])
                    # email.content_subtype = 'html'

                    # file = open("README.md", "r")
                    # email.attach("README.md", file.read(), 'text/plain')
                    email.attach_file(filename)

                    email.send()
                return render(request, "transcript2.html", dict(context, **transcriptContext))

            context = {
                'option': 1,
                'dupTranscripts': request.POST.get('dupTranscripts'),
                'form': form,
            }

            # attachment = open(filename, "rb")

            # send_mail('Transcript soft copy ready', 'Greetings, the soft copy of your transcript is ready! PFA',
            #           settings.EMAIL_HOST_USER, [toaddr], fail_silently=False)

            newUrl = '/transcript2/'+transcriptform.corrId
            return redirect(newUrl)
        else:
            context = {
                'option': 2,
                'form': form
            }
            print(form.errors)

    return render(request, "transcript.html", dict(context, **transcriptContext))


def transcriptContinuingPrice(request):
    with open('ContinueStudent.pdf', 'rb') as pdf:
        response = HttpResponse(pdf.read())
        response['content_type'] = 'application/pdf'
        response['Content-Disposition'] = 'attachment;filename=continuing-pricing-policy.pdf'
        return response


def transcriptGraduatedPrice(request):
    with open('GraduateStudent.pdf', 'rb') as pdf:
        response = HttpResponse(pdf.read())
        response['content_type'] = 'application/pdf'
        response['Content-Disposition'] = 'attachment;filename=graduated-pricing-policy.pdf'
        return response


gradesheetsDict = {}


def gradesheet(request):
    form = GradesheetForm()
    context = {
        'option': 0,
        'form': form
    }

    gradesheetContext = {
        'gradesheets': Gradesheet.objects.all(),
    }

    if request.POST:
        form = GradesheetForm(request.POST)
        # if (request.POST.get('approved')):
        #     print('***** APPROVED *****')
        if form.is_valid():

            gradesheetform = form.save(commit=False)
            gradesheetform.corrId = request.POST.get('bitsID')
            gradesheetform.corrname = request.POST.get('name')
            gradesheetform.corrhostel = request.POST.get('hostel')
            gradesheetform.corrroomNo = request.POST.get('roomNo')
            gradesheetform.ps2Station = request.POST.get('ps2Station')
            gradesheetform.email = request.POST.get('corrEmail')
            gradesheetform.phone_number = request.POST.get('phone_number')
            gradesheetform.sealedCopies = request.POST.get('sealedCopies')
            gradesheetform.postalAddress = request.POST.get('postalAddress')
            gradesheetform.postalAddressLocation = request.POST.get(
                'postalAddressLocation')
            gradesheetform.noOfSemesters = request.POST.get('noOfSemesters')

            gradesheetsDict[gradesheetform.corrId] = gradesheetform
            # gradesheetform.save()

            context = {
                'option': 1,
                'form': form,
            }
            newUrl = '/gradesheet2/'+gradesheetform.corrId
            return redirect(newUrl)
        else:
            context = {
                'option': 2,
                'form': form
            }
            print(form.errors)

    return render(request, "gradesheet.html", dict(context, **gradesheetContext))


formsDict = {}


def gradesheet2(request, bits_id):

    if bits_id not in gradesheetsDict:
        return redirect('/gradesheet')

    gradesheetform = gradesheetsDict.get(bits_id)

    no_addr = gradesheetform.sealedCopies
    no_sems = gradesheetform.noOfSemesters

    gradesheetform.amtcal = 0

    if (int(gradesheetform.sealedCopies) > 0):
        gradesheetform.amtcal += 40 + (20*int(gradesheetform.sealedCopies))

    if (gradesheetform.postalAddressLocation == 'Within India'):
        gradesheetform.amtcal += 100
    elif (gradesheetform.postalAddressLocation == 'Outside India'):
        gradesheetform.amtcal += 3000

    form = GradesheetForm2(extra=str(no_addr), sems=str(no_sems))

    context = {
        'option': 0,
        'form': form
    }

    gradesheetContext = {
        'gradesheets': Gradesheet.objects.all(),
    }

    if request.POST:
        form = GradesheetForm2(
            request.POST, extra=str(no_addr), sems=str(no_sems))

        if form.is_valid():
            formsDict[bits_id] = []
            for i in range(int(no_addr)):
                addrForm = GradesheetAddressForm()
                addrForm = addrForm.save(commit=False)
                addrForm.addr = request.POST.get('Sealed address %s' % i)
                formsDict[bits_id].append(addrForm)
                # addrForm.gradesheet = Gradesheet.objects.get(form_no = gradesheetform.form_no)
                # addrForm.save()

            for i in range(int(no_sems)):
                copyForm = GradesheetCopiesForm()
                copyForm = copyForm.save(commit=False)
                copyForm.year = request.POST.get('Year %s' % i)
                copyForm.semester = request.POST.get('Semester %s' % i)
                copyForm.copies = request.POST.get('Copies %s' % i)
                gradesheetform.amtcal += 100*int(copyForm.copies)
                formsDict[bits_id].append(copyForm)
                # copyForm.gradesheet = Gradesheet.objects.get(form_no = gradesheetform.form_no)
                # copyForm.save()

            context = {
                'option': 1,
                'form': gradesheetform,
            }
            gradesheetsDict[bits_id] = gradesheetform
            newUrl = '/gradesheet3/'+gradesheetform.corrId
            return redirect(newUrl)
        else:
            context = {
                'option': 2,
                'form': form
            }
            print(form.errors)

    return render(request, "gradesheet2.html", dict(context, **gradesheetContext))


def gradesheet3(request, bits_id):

    if bits_id not in gradesheetsDict:
        return redirect('/gradesheet')

    if bits_id not in formsDict:
        return redirect('/gradesheet')

    gradesheetform = gradesheetsDict.get(bits_id)

    form = GradesheetForm3()

    context = {
        'option': 0,
        'form': form,
        'amtcal': gradesheetform.amtcal
    }

    gradesheetContext = {
        'gradesheets': Gradesheet.objects.all(),
    }

    if request.POST:
        form = GradesheetForm3(request.POST)

        if form.is_valid():

            gradesheetform.refNo = request.POST.get('refNo')
            gradesheetform.remarks = request.POST.get('remarks')

            context = {
                'option': 1,
                'refNo': gradesheetform.refNo,
                'form': gradesheetform,
            }
            gradesheetform.save()

            for form in formsDict[bits_id]:
                form.gradesheet = Gradesheet.objects.get(
                    form_no=gradesheetform.form_no)
                form.save()

            del gradesheetsDict[bits_id]
            del formsDict[bits_id]
        else:
            context = {
                'option': 2,
                'form': form
            }
            print(form.errors)

    return render(request, "gradesheet3.html", dict(context, **gradesheetContext))
